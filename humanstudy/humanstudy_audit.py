#!/usr/bin/env python3
"""Audit human study xlsx files and generate markdown aggregate report."""

from __future__ import annotations

import argparse
import math
import statistics
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


@dataclass
class ScoreRecord:
    rater: str
    section: str
    case: str
    model: str
    score: float


@dataclass
class FileAudit:
    rater: str
    path: Path
    models: List[str] = field(default_factory=list)
    sections: List[str] = field(default_factory=list)
    valid_cases: int = 0
    valid_scores: int = 0
    missing_cells: int = 0
    out_of_range_cells: int = 0
    non_numeric_cells: int = 0
    warnings: List[str] = field(default_factory=list)


def is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def parse_models(header: Sequence[object]) -> List[str]:
    models: List[str] = []
    for cell in header[1:]:
        if cell is None:
            continue
        label = str(cell).strip()
        if not label:
            continue
        models.append(label)
    return models


def parse_file(path: Path) -> Tuple[FileAudit, List[ScoreRecord]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, ws.max_row + 1)
    ]

    audit = FileAudit(rater=path.stem, path=path)
    if not rows:
        audit.warnings.append("工作表为空。")
        return audit, []

    audit.models = parse_models(rows[0])
    if len(audit.models) == 0:
        audit.warnings.append("未识别到模型列。")
        return audit, []
    if len(audit.models) != 7:
        audit.warnings.append(f"识别到模型数量={len(audit.models)}，预期=7。")

    records: List[ScoreRecord] = []
    current_section: Optional[str] = None
    section_case_idx: Dict[str, int] = defaultdict(int)

    for row in rows[1:]:
        first = row[0] if len(row) >= 1 else None
        score_cells = list(row[1 : 1 + len(audit.models)])

        if isinstance(first, str) and first.strip():
            current_section = first.strip()
            if current_section not in audit.sections:
                audit.sections.append(current_section)

        numeric_cells = [v for v in score_cells if is_number(v)]
        # 案例行：当前有 section，且至少 3 个模型有数值
        if current_section is None or len(numeric_cells) < 3:
            continue

        section_case_idx[current_section] += 1
        audit.valid_cases += 1
        case = f"{current_section}#{section_case_idx[current_section]}"

        for model, value in zip(audit.models, score_cells):
            if value is None:
                audit.missing_cells += 1
                continue

            if not is_number(value):
                audit.non_numeric_cells += 1
                continue

            score = float(value)
            if score < 1 or score > 10:
                audit.out_of_range_cells += 1
                continue

            records.append(
                ScoreRecord(
                    rater=audit.rater,
                    section=current_section,
                    case=case,
                    model=model,
                    score=score,
                )
            )
            audit.valid_scores += 1

    if len(audit.sections) != 5:
        audit.warnings.append(f"识别到 section 数量={len(audit.sections)}，预期=5。")
    if audit.valid_cases < 15:
        audit.warnings.append(f"有效案例数量={audit.valid_cases}，低于预期约15。")

    return audit, records


def mean_std(values: Sequence[float]) -> Tuple[float, float]:
    if not values:
        return math.nan, math.nan
    mean = sum(values) / len(values)
    std = statistics.pstdev(values) if len(values) > 1 else 0.0
    return mean, std


def write_overall_ranking_csv(
    overall_rows: Sequence[Tuple[str, float, float, int]], output_csv: Path
) -> None:
    lines = ["rank,model,mean,std,n"]
    for idx, (model, mean, std, n) in enumerate(overall_rows, start=1):
        lines.append(f"{idx},{model},{mean:.6f},{std:.6f},{n}")
    output_csv.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generate_report(folder: Path, output_md: Path, output_csv: Path) -> str:
    xlsx_files = sorted(folder.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No xlsx files found in: {folder}")

    audits: List[FileAudit] = []
    records: List[ScoreRecord] = []

    for path in xlsx_files:
        audit, recs = parse_file(path)
        audits.append(audit)
        records.extend(recs)

    all_models = sorted({r.model for r in records})
    all_sections = sorted({r.section for r in records})
    all_raters = sorted({r.rater for r in records})

    model_scores: Dict[str, List[float]] = defaultdict(list)
    section_model_scores: Dict[Tuple[str, str], List[float]] = defaultdict(list)
    rater_model_scores: Dict[Tuple[str, str], List[float]] = defaultdict(list)
    section_case_count: Dict[Tuple[str, str], int] = defaultdict(int)

    for rec in records:
        model_scores[rec.model].append(rec.score)
        section_model_scores[(rec.section, rec.model)].append(rec.score)
        rater_model_scores[(rec.rater, rec.model)].append(rec.score)
        section_case_count[(rec.rater, rec.section)] = len(
            {
                r.case
                for r in records
                if r.rater == rec.rater and r.section == rec.section
            }
        )

    overall_rows = []
    for model in all_models:
        vals = model_scores[model]
        m, s = mean_std(vals)
        overall_rows.append((model, m, s, len(vals)))
    overall_rows.sort(key=lambda x: x[1])  # 分数越低越好

    lines: List[str] = []
    lines.append("# Human Study Aggregation Report")
    lines.append("")
    lines.append("## 1) 审计范围与规则")
    lines.append("")
    lines.append(f"- 扫描目录: `{folder}`")
    lines.append(f"- 发现 xlsx 文件数: `{len(xlsx_files)}`")
    lines.append("- 评分方向: `1 为最好，10 为最差（越低越好）`")
    lines.append("- 案例行识别规则: 在某个 section 下，单行至少 3 个模型存在数值")
    lines.append("- 有效分数规则: 数值且落在 `[1,10]`")
    lines.append("- 缺失值处理: 不填补，按有效样本统计")
    lines.append("")

    lines.append("## 2) 文件级审计结果")
    lines.append("")
    lines.append(
        "| rater | file | models | sections | valid_cases | valid_scores | "
        "missing_cells | non_numeric | out_of_range |"
    )
    lines.append("|---|---|---:|---:|---:|---:|---:|---:|---:|")
    for a in sorted(audits, key=lambda x: x.rater):
        lines.append(
            f"| {a.rater} | {a.path.name} | {len(a.models)} | {len(a.sections)} | "
            f"{a.valid_cases} | {a.valid_scores} | {a.missing_cells} | "
            f"{a.non_numeric_cells} | {a.out_of_range_cells} |"
        )
    lines.append("")

    warnings = [f"- `{a.rater}`: {w}" for a in audits for w in a.warnings]
    lines.append("### 审计警告")
    lines.append("")
    if warnings:
        lines.extend(warnings)
    else:
        lines.append("- 无结构性警告。")
    lines.append("")

    lines.append("## 3) 总体模型排名（全量汇总）")
    lines.append("")
    lines.append("| rank | model | mean | std | n |")
    lines.append("|---:|---|---:|---:|---:|")
    for idx, (model, mean, std, n) in enumerate(overall_rows, start=1):
        lines.append(f"| {idx} | {model} | {mean:.3f} | {std:.3f} | {n} |")
    lines.append("")

    lines.append("## 4) 分维度排名（Section-wise）")
    lines.append("")
    for section in all_sections:
        lines.append(f"### {section}")
        lines.append("")
        section_rows = []
        for model in all_models:
            vals = section_model_scores[(section, model)]
            if not vals:
                continue
            m, _ = mean_std(vals)
            section_rows.append((model, m, len(vals)))
        section_rows.sort(key=lambda x: x[1])

        lines.append("| rank | model | mean | n |")
        lines.append("|---:|---|---:|---:|")
        for idx, (model, mean, n) in enumerate(section_rows, start=1):
            lines.append(f"| {idx} | {model} | {mean:.3f} | {n} |")
        lines.append("")

    lines.append("## 5) 评分者一致性参考（每位评分者的模型均值）")
    lines.append("")
    for rater in all_raters:
        lines.append(f"### {rater}")
        lines.append("")
        rater_rows = []
        for model in all_models:
            vals = rater_model_scores[(rater, model)]
            if not vals:
                continue
            m, _ = mean_std(vals)
            rater_rows.append((model, m, len(vals)))
        rater_rows.sort(key=lambda x: x[1])
        lines.append("| rank | model | mean | n |")
        lines.append("|---:|---|---:|---:|")
        for idx, (model, mean, n) in enumerate(rater_rows, start=1):
            lines.append(f"| {idx} | {model} | {mean:.3f} | {n} |")
        lines.append("")

    lines.append("## 6) Section 完整性（每位评分者每个 Section 的案例数）")
    lines.append("")
    lines.append("| rater | section | cases |")
    lines.append("|---|---|---:|")
    for rater in all_raters:
        for section in all_sections:
            lines.append(
                f"| {rater} | {section} | {section_case_count[(rater, section)]} |"
            )
    lines.append("")

    output_md.write_text("\n".join(lines), encoding="utf-8")
    write_overall_ranking_csv(overall_rows, output_csv)
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Audit all xlsx files in folder and generate markdown report."
    )
    parser.add_argument(
        "--folder",
        type=Path,
        default=Path(__file__).resolve().parent,
        help="Folder containing xlsx files (default: script directory).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent / "humanstudy_aggregate_report.md",
        help="Output markdown path.",
    )
    parser.add_argument(
        "--output-csv",
        type=Path,
        default=Path(__file__).resolve().parent / "humanstudy_model_overall_ranking.csv",
        help="Output overall ranking csv path.",
    )
    args = parser.parse_args()

    report = generate_report(args.folder, args.output, args.output_csv)
    print(f"[OK] Report generated: {args.output}")
    print(f"[OK] CSV generated: {args.output_csv}")
    print(f"[INFO] Length: {len(report.splitlines())} lines")


if __name__ == "__main__":
    main()
